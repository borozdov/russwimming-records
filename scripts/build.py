#!/usr/bin/env python3
"""Build the public/ site from data/russia.json.

Outputs:
  public/index.html         — main page
  public/assets/style.css   — copied from static/
  public/assets/app.js      — copied from static/
  public/records.json       — same as data/russia.json (canonical API copy)
  public/records.csv        — all records in one CSV
  public/records.xlsx       — multi-sheet workbook (one sheet per category)
  public/records.md         — markdown tables
  public/records.txt        — fixed-width plain text
"""
from __future__ import annotations

import csv
import html
import json
import shutil
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent.parent
DATA = ROOT / "data" / "russia.json"
STATIC = ROOT / "static"
PUBLIC = ROOT / "public"
ASSETS = PUBLIC / "assets"

SITE_TITLE = "Рекорды России по плаванию"
SITE_TAGLINE = "Актуальные рекорды России по плаванию — вольный стиль, брасс, баттерфляй, спина, комплекс. Мужчины и женщины, бассейны 50 м и 25 м."
SITE_KEYWORDS = "рекорды России по плаванию, плавание рекорды, вольный стиль, брасс, баттерфляй, на спине, комплексное плавание, бассейн 50м, бассейн 25м, russwimming"
SITE_DOMAIN = "russwimming-records.borozdov.ru"
REPO_URL_ENV = "REPO_URL"  # optional, set in workflow; shown as "История" link

CSV_HEADERS = [
    "Категория", "Дисциплина", "Тип", "Спортсмен / Команда", "Состав эстафеты",
    "Результат", "Результат (сек)", "Место", "Дата", "Дата (ISO)",
]


def load_data() -> dict:
    return json.loads(DATA.read_text(encoding="utf-8"))


def record_rows_for_csv(data: dict):
    for cat in data["categories"]:
        for r in cat["records"]:
            yield [
                cat["title"],
                r["discipline"],
                "эстафета" if r["relay"] else "личное",
                r["athlete"],
                " · ".join(r["roster"]) if r["roster"] else "",
                r["result"],
                f"{r['result_seconds']:.2f}" if r["result_seconds"] is not None else "",
                r["location"],
                r["date_original"],
                r["date"] or "",
            ]


def write_csv(data: dict, out: Path) -> None:
    with out.open("w", encoding="utf-8-sig", newline="") as f:
        w = csv.writer(f, dialect="excel")
        w.writerow(CSV_HEADERS)
        for row in record_rows_for_csv(data):
            w.writerow(row)


def write_xlsx(data: dict, out: Path) -> None:
    wb = Workbook()
    wb.remove(wb.active)

    header_font = Font(bold=True, color="FAFAFA")
    header_fill = PatternFill("solid", fgColor="0D0D0D")
    mono_font = Font(name="Menlo")
    center = Alignment(horizontal="center", vertical="center")

    for cat in data["categories"]:
        title = cat["title"][:31]  # sheet name limit
        ws = wb.create_sheet(title=title)
        headers = ["Дисциплина", "Спортсмен / Команда", "Состав эстафеты",
                   "Результат", "Место", "Дата"]
        ws.append(headers)
        for c in ws[1]:
            c.font = header_font
            c.fill = header_fill
            c.alignment = center
        for r in cat["records"]:
            ws.append([
                r["discipline"],
                r["athlete"],
                " · ".join(r["roster"]) if r["roster"] else "",
                r["result"],
                r["location"],
                r["date_original"],
            ])
        # column widths
        widths = [34, 26, 42, 12, 22, 14]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w
        # monospace for the result column (col 4)
        for row_idx in range(2, ws.max_row + 1):
            ws.cell(row=row_idx, column=4).font = mono_font
        ws.freeze_panes = "A2"

    # Summary / all-in-one sheet
    ws = wb.create_sheet(title="Все", index=0)
    ws.append(CSV_HEADERS)
    for c in ws[1]:
        c.font = header_font
        c.fill = header_fill
        c.alignment = center
    for row in record_rows_for_csv(data):
        ws.append(row)
    widths = [28, 34, 10, 24, 42, 12, 14, 22, 14, 14]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"

    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)


def write_markdown(data: dict, out: Path) -> None:
    lines = [
        f"# {SITE_TITLE}",
        "",
        f"Источник: <{data['source_url']}>  ",
        f"Обновлено: {data['fetched_at']}  ",
        f"Всего рекордов: {data['total_records']}",
        "",
    ]
    for cat in data["categories"]:
        lines.append(f"## {cat['title']}")
        lines.append("")
        lines.append("| Дисциплина | Спортсмен | Результат | Место | Дата |")
        lines.append("|---|---|---|---|---|")
        for r in cat["records"]:
            athlete = r["athlete"]
            if r["roster"]:
                athlete += " (" + ", ".join(r["roster"]) + ")"
            lines.append(
                f"| {r['discipline']} | {athlete} | `{r['result']}` | "
                f"{r['location']} | {r['date_original']} |"
            )
        lines.append("")
    out.write_text("\n".join(lines) + "\n", encoding="utf-8")


def write_txt(data: dict, out: Path) -> None:
    """Fixed-width plain-text dump — handy for grep / terminal users."""
    rows = []
    for cat in data["categories"]:
        rows.append(f"=== {cat['title']} ===")
        for r in cat["records"]:
            athlete = r["athlete"]
            if r["roster"]:
                athlete += " (" + ", ".join(r["roster"]) + ")"
            rows.append(f"  {r['discipline']:<42}  {r['result']:>9}  {athlete:<60}  {r['location']:<22}  {r['date_original']}")
        rows.append("")
    out.write_text("\n".join(rows) + "\n", encoding="utf-8")


def copy_static() -> None:
    ASSETS.mkdir(parents=True, exist_ok=True)
    for name in ("style.css", "app.js"):
        shutil.copy2(STATIC / name, ASSETS / name)


def fmt_date_ru(iso: str, orig: str) -> str:
    return ".".join(reversed(iso.split("-"))) if iso else (orig or "")


def latest_record_date(data: dict) -> str:
    dates = [r["date"] for c in data["categories"] for r in c["records"] if r["date"]]
    return ".".join(reversed(max(dates).split("-"))) if dates else "—"


def render_table_html(data: dict) -> str:
    """Серверный рендер таблицы: контент виден поисковикам и без JS.
    Разметка идентична клиентскому рендеру в static/app.js (состояние по умолчанию)."""
    fresh_year = data["fetched_at"][:4]
    e = html.escape

    def badges(r: dict) -> str:
        out = ""
        if r["date"] and r["date"].startswith(fresh_year):
            out += '<span class="badge badge-fresh">Новое</span>'
        if r["relay"]:
            out += '<span class="badge badge-relay">Эстафета</span>'
        return out

    cols = [
        ("discipline", "Дисциплина", False),
        ("athlete", "Спортсмен", False),
        ("result", "Результат", True),
        ("location", "Место", False),
        ("date", "Дата", True),
    ]
    head = "".join(
        f'<th class="{"num" if num else ""}" data-key="{key}" aria-sort="none" '
        f'title="Сортировать">{label}<span class="sort-ind"></span></th>'
        for key, label, num in cols
    )

    rows, cards = [], []
    for cat in data["categories"]:
        rows.append(
            f'<tr class="group-row"><td colspan="5">{e(cat["title"])}'
            f'<span class="group-count">{len(cat["records"])}</span></td></tr>'
        )
        cards.append(f'<div class="card-group-head">{e(cat["title"])}</div>')
        for r in cat["records"]:
            b = badges(r)
            date_h = fmt_date_ru(r["date"], r["date_original"])
            roster = f'<div class="roster">{e(" · ".join(r["roster"]))}</div>' if r["roster"] else ""
            rows.append(
                "<tr>"
                f'<td class="col-disc">{e(r["discipline"])}{b}</td>'
                f'<td>{e(r["athlete"])}{roster}</td>'
                f'<td class="col-result">{e(r["result"])}</td>'
                f'<td class="col-loc">{e(r["location"])}</td>'
                f'<td class="col-date">{date_h}</td>'
                "</tr>"
            )
            roster_c = f' · {e(", ".join(r["roster"]))}' if r["roster"] else ""
            cards.append(
                '<article class="card">'
                '<div class="card-top">'
                f'<span class="card-disc">{e(r["discipline"])}{b}</span>'
                f'<span class="card-result">{e(r["result"])}</span>'
                "</div>"
                f'<div class="card-athlete">{e(r["athlete"])}{roster_c}</div>'
                f'<div class="card-meta">{e(r["location"])} · <span class="mono">{date_h}</span></div>'
                "</article>"
            )

    return (
        '<div class="table-frame">'
        f'<div class="table-scroll"><table class="records"><thead><tr>{head}</tr></thead>'
        f'<tbody>{"".join(rows)}</tbody></table></div>'
        f'<div class="cards">{"".join(cards)}</div>'
        "</div>"
    )


FAVICON = (
    "data:image/svg+xml;utf8,"
    "<svg xmlns='http://www.w3.org/2000/svg' viewBox='0 0 100 100'>"
    "<rect width='100' height='100' rx='16' fill='%23fafafa'/>"
    "<text x='50' y='53' dominant-baseline='central' text-anchor='middle' "
    "font-family='Inter,Helvetica,Arial,sans-serif' font-weight='700' font-size='60' "
    "fill='%230d0d0d'>Р</text></svg>"
)

INDEX_TEMPLATE = """<!doctype html>
<html lang="ru" data-theme="obsidian">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<script>
// Лик следует за средой; ручной выбор — с памятью (брендбук §1.5)
(function () {{
  try {{
    var s = localStorage.getItem("lik");
    var lik = (s === "titan" || s === "obsidian") ? s :
      (window.matchMedia("(prefers-color-scheme: light)").matches ? "titan" : "obsidian");
    document.documentElement.setAttribute("data-theme", lik);
  }} catch (e) {{}}
}})();
</script>
<title>{title}</title>
<meta name="description" content="{description}">
<meta name="keywords" content="{keywords}">
<meta name="robots" content="index, follow">
<link rel="canonical" href="https://{domain}/">
<meta property="og:type" content="website">
<meta property="og:site_name" content="Рекорды России по плаванию">
<meta property="og:title" content="{title}">
<meta property="og:description" content="{description}">
<meta property="og:url" content="https://{domain}/">
<meta property="og:locale" content="ru_RU">
<meta name="twitter:card" content="summary">
<meta name="twitter:title" content="{title}">
<meta name="twitter:description" content="{description}">
<meta name="theme-color" media="(prefers-color-scheme: dark)" content="#0d0d0d">
<meta name="theme-color" media="(prefers-color-scheme: light)" content="#fafafa">
<meta name="google-site-verification" content="qOwWmdq24kGcVTyxc1GL2W8TxQk63Z5lBH3NSv4hH4s" />
<meta name="yandex-verification" content="80f947e774535d84" />
<link rel="icon" href="{favicon}">
<link rel="preconnect" href="https://fonts.googleapis.com">
<link rel="preconnect" href="https://fonts.gstatic.com" crossorigin>
<link href="https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700&family=JetBrains+Mono:wght@400;500;700&display=swap" rel="stylesheet">
<link rel="stylesheet" href="./assets/style.css">
<!-- Yandex.Metrika counter -->
<script type="text/javascript">
    (function(m,e,t,r,i,k,a){{
        m[i]=m[i]||function(){{(m[i].a=m[i].a||[]).push(arguments)}};
        m[i].l=1*new Date();
        for (var j = 0; j < document.scripts.length; j++) {{if (document.scripts[j].src === r) {{ return; }}}}
        k=e.createElement(t),a=e.getElementsByTagName(t)[0],k.async=1,k.src=r,a.parentNode.insertBefore(k,a)
    }})(window, document,'script','https://mc.yandex.ru/metrika/tag.js?id=109049034', 'ym');

    ym(109049034, 'init', {{ssr:true, webvisor:true, clickmap:true, ecommerce:"dataLayer", referrer: document.referrer, url: location.href, accurateTrackBounce:true, trackLinks:true}});
</script>
<noscript><div><img src="https://mc.yandex.ru/watch/109049034" style="position:absolute; left:-9999px;" alt="" /></div></noscript>
<!-- /Yandex.Metrika counter -->
<script type="application/ld+json">{jsonld}</script>
</head>
<body>
<a class="skip-link" href="#table">К таблице рекордов</a>

<header class="site-header">
  <div class="shell header-inner">
    <a class="wordmark" href="/">Рекорды России<span class="wordmark-sub">Плавание</span></a>
    <div class="header-actions">
      <button class="icon-btn" id="theme-toggle" aria-label="Сменить лик"><svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round" aria-hidden="true"><circle cx="12" cy="12" r="4"/><path d="M12 2v2M12 20v2M4.93 4.93l1.41 1.41M17.66 17.66l1.41 1.41M2 12h2M20 12h2M4.93 19.07l1.41-1.41M17.66 6.34l1.41-1.41"/></svg></button>
      <button class="btn btn-print" id="print-btn">Печать</button>
      <div class="dl" id="dl-menu">
        <button class="btn btn-primary" id="dl-btn" aria-haspopup="true" aria-expanded="false">
          <svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round" aria-hidden="true"><path d="M12 3v12M6 11l6 6 6-6M4 21h16"/></svg>
          Скачать
        </button>
        <div class="dl-panel">
          <a href="./records.json" download><span class="fmt">JSON</span><span class="hint">структурированные данные</span></a>
          <a href="./records.csv" download><span class="fmt">CSV</span><span class="hint">Excel / Numbers</span></a>
          <a href="./records.xlsx" download><span class="fmt">XLSX</span><span class="hint">книга по листам</span></a>
          <a href="./records.md" download><span class="fmt">MD</span><span class="hint">таблицы Markdown</span></a>
          <a href="./records.txt" download><span class="fmt">TXT</span><span class="hint">фиксированная ширина</span></a>
          <a id="dl-png-btn" href="#"><span class="fmt">PNG</span><span class="hint">картинка таблицы</span></a>
          <a id="dl-pdf-btn" href="#"><span class="fmt">PDF</span><span class="hint">документ таблицы</span></a>
        </div>
      </div>
    </div>
  </div>
</header>

<main class="shell">
  <section class="hero">
    <p class="label">Официальное зеркало · Всероссийская федерация плавания</p>
    <h1>Рекорды России по&nbsp;плаванию</h1>
    <p class="hero-meta">
      Действующие рекорды России: вольный стиль, на спине, брасс, баттерфляй,
      комплекс и эстафеты. Мужчины, женщины и смешанные команды, бассейны 50 и 25 метров.
    </p>
    <p class="hero-links">
      Источник: <a href="{source_url}" rel="noopener" target="_blank">russwimming.ru/records/russia/</a>
      · Обновляется автоматически раз в сутки
    </p>
    <div class="stat-strip">
      <div class="stat">
        <div class="stat-value">{total}</div>
        <div class="stat-note"><span class="stat-note-text">Действующих рекордов</span><span class="label">В таблице</span></div>
      </div>
      <div class="stat">
        <div class="stat-value">{latest_record}</div>
        <div class="stat-note"><span class="stat-note-text">Последний рекорд</span><span class="label">Дата</span></div>
      </div>
      <div class="stat">
        <div class="stat-value"><time datetime="{fetched_at_iso}">{fetched_at_human}</time></div>
        <div class="stat-note"><span class="stat-note-text">Данные обновлены</span><span class="label">Раз в сутки</span></div>
      </div>
    </div>
  </section>

  <section class="controls" aria-label="Поиск и фильтры">
    <div class="search-row">
      <label class="search">
        <svg class="search-icon" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" aria-hidden="true"><circle cx="11" cy="11" r="7"/><path d="m21 21-4.3-4.3"/></svg>
        <input id="search" type="search" placeholder="Фамилия, дисциплина, город…" autocomplete="off" aria-label="Поиск по таблице">
        <span class="search-key" aria-hidden="true">/</span>
      </label>
      <span class="result-count label">Показано <span id="visible-count"><b class="mono">{total}</b> из <span class="mono">{total}</span></span></span>
    </div>
    <div class="filters" id="filters" aria-label="Фильтры"></div>
  </section>

  <div id="table">{table_html}</div>

  <footer class="footer">
    <div class="footer-inner">
      <div>
        <p>
          Данные синхронизируются раз в сутки с
          <a href="{source_url}" rel="noopener" target="_blank">russwimming.ru</a>.
          Если рекорд отсутствует или выглядит устаревшим — сайт-источник ещё не обновил свою таблицу.
        </p>
        <p>
          Вопросы и предложения: <a href="https://t.me/BorozdovNikita" rel="noopener" target="_blank">@BorozdovNikita</a>
          {history_link}
        </p>
      </div>
      <div class="signature">By <a href="https://borozdov.ru" rel="noopener" target="_blank">Borozdov</a></div>
    </div>
    <div class="print-signature"><span class="print-date"></span></div>
  </footer>
</main>

<div class="toast" id="toast" role="status" aria-live="polite"></div>

<script id="records-data" type="application/json">{data_json}</script>
<script src="./assets/app.js"></script>
</body>
</html>
"""


def write_index(data: dict, out: Path) -> None:
    fetched_dt = datetime.strptime(data["fetched_at"], "%Y-%m-%dT%H:%M:%SZ").replace(tzinfo=timezone.utc)
    fetched_human = fetched_dt.strftime("%d.%m.%Y")

    import os
    repo_url = os.environ.get(REPO_URL_ENV, "").strip()
    history_link = ""
    if repo_url:
        repo_url = repo_url.rstrip("/")
        history_link = (
            f'· <a href="{repo_url}/commits/main/data/russia.json" '
            f'rel="noopener" target="_blank">история изменений</a>'
        )

    jsonld = json.dumps({
        "@context": "https://schema.org",
        "@type": "Dataset",
        "name": SITE_TITLE,
        "description": SITE_TAGLINE,
        "url": f"https://{SITE_DOMAIN}/",
        "dateModified": data["fetched_at"],
        "license": "https://creativecommons.org/licenses/by/4.0/",
        "creator": {"@type": "Organization", "name": "Всероссийская федерация плавания", "url": "https://russwimming.ru"},
        "distribution": [
            {"@type": "DataDownload", "encodingFormat": "application/json", "contentUrl": f"https://{SITE_DOMAIN}/records.json"},
            {"@type": "DataDownload", "encodingFormat": "text/csv", "contentUrl": f"https://{SITE_DOMAIN}/records.csv"},
        ],
    }, ensure_ascii=False)

    html_doc = INDEX_TEMPLATE.format(
        title=html.escape(SITE_TITLE),
        description=html.escape(SITE_TAGLINE),
        keywords=html.escape(SITE_KEYWORDS),
        domain=SITE_DOMAIN,
        favicon=FAVICON,
        jsonld=jsonld,
        source_url=html.escape(data["source_url"]),
        total=data["total_records"],
        latest_record=latest_record_date(data),
        fetched_at_iso=data["fetched_at"],
        fetched_at_human=fetched_human,
        history_link=history_link,
        table_html=render_table_html(data),
        data_json=json.dumps(data, ensure_ascii=False).replace("</", "<\\/"),
    )
    out.write_text(html_doc, encoding="utf-8")


ROBOTS_TXT = f"""User-agent: *
Allow: /

Sitemap: https://{SITE_DOMAIN}/sitemap.xml
"""

NOT_FOUND_HTML = f"""<!doctype html>
<html lang="ru">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>Страница не найдена — Рекорды России по плаванию</title>
<meta http-equiv="refresh" content="0; url=https://{SITE_DOMAIN}/">
<link rel="canonical" href="https://{SITE_DOMAIN}/">
</head>
<body>
<p>Страница не найдена. <a href="https://{SITE_DOMAIN}/">Перейти на главную →</a></p>
<script>window.location.replace("https://{SITE_DOMAIN}/")</script>
</body>
</html>
"""


def write_sitemap(data: dict, out: Path) -> None:
    lastmod = data["fetched_at"][:10]
    sitemap = (
        '<?xml version="1.0" encoding="UTF-8"?>\n'
        '<urlset xmlns="http://www.sitemaps.org/schemas/sitemap/0.9">\n'
        f'  <url>\n'
        f'    <loc>https://{SITE_DOMAIN}/</loc>\n'
        f'    <lastmod>{lastmod}</lastmod>\n'
        f'    <changefreq>daily</changefreq>\n'
        f'    <priority>1.0</priority>\n'
        f'  </url>\n'
        '</urlset>\n'
    )
    out.write_text(sitemap, encoding="utf-8")


def main() -> int:
    data = load_data()
    PUBLIC.mkdir(parents=True, exist_ok=True)
    ASSETS.mkdir(parents=True, exist_ok=True)

    copy_static()
    write_index(data, PUBLIC / "index.html")
    write_sitemap(data, PUBLIC / "sitemap.xml")
    (PUBLIC / "robots.txt").write_text(ROBOTS_TXT, encoding="utf-8")
    (PUBLIC / "404.html").write_text(NOT_FOUND_HTML, encoding="utf-8")
    # Public copy of the canonical data
    (PUBLIC / "records.json").write_text(
        json.dumps(data, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )
    write_csv(data, PUBLIC / "records.csv")
    write_xlsx(data, PUBLIC / "records.xlsx")
    write_markdown(data, PUBLIC / "records.md")
    write_txt(data, PUBLIC / "records.txt")

    print(f"built {PUBLIC.relative_to(ROOT)}/ — {data['total_records']} records")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
